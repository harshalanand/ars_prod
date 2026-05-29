"""Generate listing_allocation_inputs_master.xlsx — the operator's pre-flight
checklist for everything that must be in place (UI knobs + DB tables) before
running Listing Generation + Allocation.

Output: backend/app/docs/reference/listing_allocation_inputs_master.xlsx
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR = Path(__file__).resolve().parent.parent / "app" / "docs" / "reference"
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT_PATH = OUT_DIR / "listing_allocation_inputs_master.xlsx"

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="305496")
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
WRAP = Alignment(wrap_text=True, vertical="top")


def write_sheet(wb, title, headers, rows, col_widths=None):
    ws = wb.create_sheet(title=title)
    # Header
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        c.border = BORDER
    # Body
    for ri, row in enumerate(rows, start=2):
        for ci, val in enumerate(row, start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = WRAP
            c.border = BORDER
    # Column widths
    if col_widths:
        for ci, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(ci)].width = w
    # Freeze header row + first column
    ws.freeze_panes = "B2"
    # Auto row height heuristic
    for ri in range(2, len(rows) + 2):
        ws.row_dimensions[ri].height = 38


# ---------------------------------------------------------------------------
# Sheet 1 — UI INPUTS (Listing Generation form)
# ---------------------------------------------------------------------------
UI_HDR = [
    "Setting", "UI label / API param", "Type", "Default", "Valid range",
    "Description (what it controls)", "Affects (Part / Stage)", "When to change",
    "Required?",
]
UI_ROWS = [
    ["Stock %", "stock_threshold_pct / size_threshold", "Number", "0.6", "0 .. 1",
     "Adequacy threshold for OPT_TYPE classification AND size-coverage ratio for R07 everywhere.",
     "Part 3.6, Part 7, Stage A.2 R07, Stage C R07-live, Part 8.5",
     "Raise to 0.7-0.8 to be stricter (more TBC/TBL). Lower for more RL.", "Yes"],
    ["MinSz", "min_size_count", "Integer", "3", "0 .. n",
     "Absolute floor on VAR_FNL_COUNT for TBL eligibility. 0 disables the count half of R07.",
     "Part 3.6 MIX(b), Part 7, Stage A.2 R07, Stage C R07-live",
     "Set to 0 to allow thin colours through; raise to flag more as MIX.", "Yes"],
    ["Default ACS_D", "default_acs_d", "Number", "18", "1 .. 100",
     "Fallback when ACS_D from ARS_CALC_ST_MAJ_CAT is 0 / NULL.",
     "Part 3.6 (threshold × ACS_D), Part 4c (OPT_MBQ)",
     "Lower (e.g. 12) for new stores with no sales history.", "Yes"],
    ["MJ_REQ growth %", "mj_req_growth_pct", "Number", "100", "100 .. 200",
     "Scales MJ_MBQ into MJ_MBQ_REV (e.g. 110 = 10% lift). Drives every downstream consumer.",
     "Part 7 (MJ_MBQ_REV / MJ_REQ_REV)", "Raise during peak season to lift targets.", "Yes"],
    ["RL MBQ cap %", "rl_mbq_cap_pct", "Number", "130", "100 .. 200",
     "RL waterfall ceiling. budget = max(0, MJ_REQ_REM + ((cap-100)/100) × MJ_MBQ).",
     "Stage C _build_mbq_budget / _live_mbq_budget for RL",
     "Lower to 100 to disable RL growth; pri_ct_check_rl=True pins to 100.", "Yes"],
    ["TBC MBQ cap %", "tbc_mbq_cap_pct", "Number", "130", "100 .. 200",
     "TBC waterfall ceiling. Same math as RL.",
     "Stage C _live_mbq_budget for TBC",
     "Symmetric with RL knob.", "Yes"],
    ["RL MJ_REQ cap %", "rl_mj_req_cap_pct", "Number", "(varies)", "0 .. 200",
     "Per-OPT-type cap on MJ_REQ headroom; participates in R09 inter-OPT_TYPE check.",
     "Stage D _check_r09_eligibility (after RL)",
     "Lower if RL is over-eating warehouse pool.", "Optional"],
    ["TBC MJ_REQ cap %", "tbc_mj_req_cap_pct", "Number", "(varies)", "0 .. 200",
     "Same as RL but applied after TBC.",
     "Stage D _check_r09_eligibility (after TBC)",
     "Symmetric.", "Optional"],
    ["TBL MJ_REQ cap %", "tbl_mj_req_cap_pct", "Number", "(varies)", "0 .. 200",
     "Same as RL/TBC but applied after TBL.",
     "Stage D _check_r09_eligibility (after TBL)",
     "Symmetric.", "Optional"],
    ["PRI strict RL", "pri_ct_check_rl", "Toggle", "False (Off)", "True / False",
     "When On: RL enforces PRI_CT% ≥ 100 (R06). Pins eff_rl_cap to 100 (no growth).",
     "Stage A.2 R06; Stage C cap selection; _pre_band_check",
     "Turn On for strict primary-coverage runs.", "Yes"],
    ["PRI strict TBC", "pri_ct_check_tbc", "Toggle", "False (Off)", "True / False",
     "Symmetric to PRI strict RL.",
     "Stage A.2 R06; Stage C cap selection",
     "Symmetric.", "Yes"],
    ["Apply Sec-cap in normal mode", "apply_sec_cap_in_normal", "Toggle", "False", "True / False",
     "Apply secondary-grid MBQ caps even when not in fallback mode.",
     "Stage D _apply_sec_grid_cap_pre_gate",
     "Turn On when sec-cap-marked grids must be enforced.", "Optional"],
    ["Allocation mode", "allocation_mode", "Dropdown", "pandas", "pandas / sequential",
     "Engine choice: pandas (default, process-pooled) or sequential (single-thread SQL).",
     "Part 8 dispatcher",
     "Use sequential only for debugging single-MAJ_CAT issues.", "Yes"],
    ["Parallel workers", "parallel_workers", "Integer", "4", "1 .. 8",
     "ProcessPoolExecutor worker count. Each worker has own GIL.",
     "Stage C subprocess pool",
     "Raise to 8 on fast hardware with many MAJ_CATs.", "Yes"],
    ["Use writer queue", "use_writer_queue", "Toggle", "False (or env default)", "True / False",
     "Route all DB writes through single writer thread (eliminates writer contention).",
     "Stage C writer thread",
     "Turn On when seeing deadlock retries in logs.", "Optional"],
    ["OPT types to run", "opt_types", "List", "['RL','TBC','TBL']", "subset",
     "Which OPT_TYPE waves to run. Omitting a type skips its wave.",
     "Stage C outer loop",
     "Useful for partial reruns.", "Optional"],
    ["MAJ_CATs filter", "majcats / only_majcats", "List", "(all)", "subset of MAJ_CATs",
     "Run only specified MAJ_CATs. Used by retry path to redo failed slices.",
     "Stage A+B + queue seed",
     "Retry path or focused run.", "Optional"],
    ["RDC mode", "rdc_mode", "Dropdown", "own", "own / cross / all",
     "How stores are matched to warehouse RDC: 'own' = home RDC; 'cross' = source-from-other RDCs.",
     "Part 1/2 store join; MSA filter",
     "Switch to 'cross' for inter-RDC sourcing.", "Yes"],
    ["RDC values", "rdc_values", "List", "(all RDCs)", "list of RDC codes",
     "Restrict to specific RDCs.",
     "Part 1/2 MSA RDC filter",
     "Optional narrowing.", "Optional"],
    ["Cross-from RDCs", "cross_from", "List", "(empty)", "list of RDC codes",
     "Source RDCs for cross-mode.",
     "Cross-RDC sourcing only",
     "Pair with rdc_mode='cross'.", "Optional"],
    ["Cross-to stores", "cross_to", "List", "(empty)", "list of WERKS",
     "Destination stores for cross-mode.",
     "Cross-RDC sourcing only",
     "Pair with rdc_mode='cross'.", "Optional"],
    ["MIX mode", "mix_mode", "Dropdown", "st_maj_rng", "each / st_maj / st_maj_rng",
     "How MIX rows are aggregated in Part 3.7: each = keep all; st_maj = roll up to (store, MAJ_CAT); st_maj_rng = roll up by RNG.",
     "Part 3.7",
     "'each' preserves individual MIX rows for audit.", "Yes"],
    ["Store ranking req_weight", "req_weight", "Number", "0.4", "0 .. 1",
     "Weight applied to REQ_RANK when computing W_SCORE for ST_RANK.",
     "Part 6 store ranking",
     "Raise to favour high-demand stores.", "Yes"],
    ["Store ranking fill_weight", "fill_weight", "Number", "0.6", "0 .. 1",
     "Weight applied to FILL_RANK when computing W_SCORE.",
     "Part 6 store ranking",
     "Raise to favour empty stores.", "Yes"],
    ["TBL trivial factor", "tbl_trivial_factor", "Number", "(see config)", "0 .. 1",
     "Reserved/legacy — currently unused. Kept in API signature.",
     "n/a",
     "Leave at default.", "Optional"],
    ["MSA table", "msa_table", "Dropdown", "ARS_MSA_TOTAL", "table name",
     "Which MSA table to use as warehouse source.",
     "Part 2, Part 3.55, Stage B.1",
     "Almost always ARS_MSA_TOTAL.", "Yes"],
    ["Grid table", "grid_table", "Dropdown", "ARS_GRID_MJ_GEN_ART", "table name",
     "Which GEN_ART-grain grid to use for Part 1.",
     "Part 1",
     "Driven by current Grid Builder output.", "Yes"],
    ["CONT table", "cont_table", "Dropdown", "Master_CONT_SZ", "table name",
     "Size contribution source.",
     "Stage B.2 _enrich_size_cont",
     "Usually Master_CONT_SZ.", "Yes"],
    ["Var-grid table", "var_grid_table", "Dropdown", "ARS_GRID_MJ_VAR_ART", "table name",
     "Variant-grain grid for explode + per-size stock lookup.",
     "Stage B.1 _stage_b_explode",
     "Driven by Grid Builder.", "Yes"],
]

# ---------------------------------------------------------------------------
# Sheet 2 — MASTER DATA TABLES (must be populated before run)
# ---------------------------------------------------------------------------
DATA_HDR = [
    "Table", "Owner / source process", "Grain (PK / unique keys)",
    "Required columns", "Used by (Part / Stage)", "How to refresh",
    "Required?", "Impact if missing / stale", "Notes",
]
DATA_ROWS = [
    # ─── MSA family
    ["ARS_MSA_TOTAL", "MSA Calculation pipeline",
     "(RDC, ARTICLE_NUMBER)",
     "RDC, MAJ_CAT, GEN_ART_NUMBER, CLR, ARTICLE_NUMBER, STK_QTY, PEND_QTY, HOLD_QTY, FNL_Q",
     "Part 2 (DISTINCT triples), apply_pend_alc_delta",
     "Run MSA Stock Calculation in UI (Data Preparation → MSA Stock Calculation)",
     "Yes",
     "Empty/missing → no IS_NEW=1 candidates; allocator has no warehouse view.",
     "Single source of truth for warehouse stock."],
    ["ARS_MSA_GEN_ART", "MSA Calculation pipeline (rollup)",
     "(RDC, MAJ_CAT, GEN_ART_NUMBER, CLR)",
     "RDC, MAJ_CAT, GEN_ART_NUMBER, CLR, PEND_QTY, HOLD_QTY, FNL_Q",
     "Part 3.55 (MSA_FNL_Q join)",
     "Auto-rolled-up from ARS_MSA_TOTAL by the MSA pipeline.",
     "Yes",
     "MSA_FNL_Q stays 0; OPT_TYPE falls to MIX(a) for every row.",
     "Auto-maintained."],
    ["ARS_MSA_VAR_ART", "MSA Calculation pipeline (variant)",
     "(RDC, MAJ_CAT, GEN_ART_NUMBER, CLR, VAR_ART)",
     "RDC, MAJ_CAT, GEN_ART_NUMBER, CLR, VAR_ART, PEND_QTY, HOLD_QTY, FNL_Q",
     "Part 3.55 (VAR_COUNT, VAR_FNL_COUNT), Stage B.1 (explode)",
     "Auto-rolled-up by MSA pipeline.",
     "Yes",
     "VAR_COUNT/FNL_COUNT = 0; R07 size-coverage cannot evaluate; TBL filter passes everything.",
     "Variant-grain pool source."],

    # ─── Grids (Grid Builder)
    ["ARS_GRID_BUILDER", "Grid Builder UI",
     "grid_name UNIQUE",
     "grid_name, hierarchy_columns, output_table, status (Active/Inactive), seq, sec_cap_applicable, sec_cap_pct, use_for_opt_sale",
     "Part 4a, _discover_primary_grids, _discover_grid_rollup_tables",
     "Configure in UI → Data Preparation → Grid Builder.",
     "Yes",
     "No active grids → no per-grid MBQ/REQ columns; Part 4e silent.",
     "Inactive grids skipped from PEND_ALC delta (recent fix)."],
    ["ARS_GRID_MJ_GEN_ART", "Grid Builder (per-run output)",
     "(WERKS, MAJ_CAT, GEN_ART_NUMBER, CLR)",
     "WERKS, MAJ_CAT, GEN_ART_NUMBER, CLR, STK_E01, STK_E02, ..., PEND_ALC",
     "Part 1 (INSERT source)",
     "Run Grid Builder → MJ_GEN_ART grid.",
     "Yes",
     "Part 1 empty → only Part 2 MSA-only rows enter the listing.",
     "Primary grid for IS_NEW=0 rows."],
    ["ARS_GRID_MJ_VAR_ART", "Grid Builder (per-run output)",
     "(WERKS, MAJ_CAT, GEN_ART_NUMBER, CLR, VAR_ART)",
     "WERKS, MAJ_CAT, GEN_ART_NUMBER, CLR, VAR_ART, SZ stock cols, PEND_ALC",
     "Stage B.1 (variant grid for explode + SZ_STK)",
     "Run Grid Builder → MJ_VAR_ART grid.",
     "Yes",
     "Stage B.1 has no SZ_STK; alloc demand inflated.",
     "Variant-grain grid; pool key source."],
    ["ARS_GRID_MJ", "Grid Builder (per-run output)",
     "(WERKS, MAJ_CAT)",
     "WERKS, MAJ_CAT, MBQ, STK_TTL, CONT, OPT_CNT, DISP_Q, REQ, PEND_ALC",
     "Part 4a (MJ_MBQ, MJ_STK_TTL, MJ_DISP_Q, MJ_CONT, ...)",
     "Run Grid Builder → MJ grid.",
     "Yes",
     "MJ_MBQ/MJ_REQ NULL → all R09 / store-broken / cap checks degrade.",
     "Primary cap source."],
    ["ARS_GRID_MJ_FAB", "Grid Builder (optional active grid)",
     "(WERKS, MAJ_CAT, FAB)",
     "WERKS, MAJ_CAT, FAB, MBQ, STK_TTL, REQ, PEND_ALC",
     "Part 4a, sec-cap pre-gate",
     "Run Grid Builder if grid is Active.",
     "Optional",
     "Sec-cap can't apply at FAB level if missing.",
     "Activate via ARS_GRID_BUILDER.status."],
    ["ARS_GRID_MJ_MICRO_MVGR / MACRO_MVGR / M_VND_CD / RNG_SEG / WEAVE_2 / M_YARN_02 / MERGE_RNG_SEG", "Grid Builder (any active secondary grid)",
     "(WERKS, MAJ_CAT, <attr>)",
     "Same family as MJ_FAB",
     "Part 4a, sec-cap pre-gate",
     "Build via Grid Builder; toggle Active/Inactive in ARS_GRID_BUILDER.",
     "Optional",
     "Inactive grids silently skipped (recent change 2026-05-XX).",
     "User-configurable; expect 4-8 active grids in production."],
    ["ARS_GRID_HIERARCHY", "Auto-maintained by Part 7",
     "(WERKS, MAJ_CAT, [+ one col per active non-article grid])",
     "WERKS, MAJ_CAT + one col per active grid's last hier col",
     "Part 7 refresh; sec-cap evaluation",
     "Auto-rebuilt on every Part 7 run.",
     "Auto",
     "Sec-cap can't join — silent drop.",
     "Managed; do not edit manually."],
    ["ARS_MERGE_RULES", "Merge Rules UI",
     "(parent_col, source_value)",
     "parent_col, source_value, target_value, is_active",
     "Grid Builder MERGE_<X> grid building, derived_masters at delta time",
     "Configure in UI → Data Preparation → Merge Rules.",
     "Optional (only when MERGE grids used)",
     "MERGE_<X> grids will fail to build / resolve.",
     "Drives MERGE_RNG_SEG and similar derived columns."],

    # ─── Calc / Contribution outputs
    ["ARS_CALC_ST_MAJ_CAT", "Contribution pipeline (Step ST_MAJ_CAT)",
     "(ST_CD, MAJ_CAT)",
     "ST_CD, MAJ_CAT, ACS_D, ALC_D, LISTING, I_ROD, CLR_MIN, CLR_MAX",
     "Part 3.5, Part 3.5a Step 1, Part 6 (W_SCORE inputs)",
     "Run Contribution pipeline before Listing.",
     "Yes",
     "Stale ACS_D → wrong OPT_MBQ; missing LISTING/I_ROD → defaults apply.",
     "Baseline per-store-per-category settings."],
    ["ARS_CALC_ST_ART", "Contribution pipeline (Step ST_ART)",
     "(ST_CD, MAJ_CAT, GEN_ART_NUMBER [, CLR])",
     "ST_CD, MAJ_CAT, GEN_ART_NUMBER, CLR, LISTING, I_ROD, FOCUS_W_CAP, FOCUS_WO_CAP",
     "Part 3.5a Step 2 (article-level overrides)",
     "Run Contribution pipeline.",
     "Optional (cascade only)",
     "Article-level overrides not applied; MAJ_CAT baseline used.",
     "Cascade source; nullable values use baseline."],
    ["MASTER_GEN_ART_SALE", "Contribution pipeline (Step master_sale_sal_pd)",
     "(ST_CD, MAJ_CAT, GEN_ART_NUMBER, CLR)",
     "ST_CD, MAJ_CAT, GEN_ART_NUMBER, CLR, SAL_PD",
     "Part 3.5b (AUTO_GEN_ART_SALE)",
     "Run Contribution pipeline.",
     "Yes",
     "AUTO_GEN_ART_SALE stays NULL; MAX_DAILY_SALE ranking degraded.",
     "Article-level sale velocity (~21L rows)."],
    ["MASTER_GEN_ART_AGE", "Contribution pipeline / external upload",
     "(ST_CD, MAJ_CAT, GEN_ART_NUMBER, CLR)",
     "ST_CD, MAJ_CAT, GEN_ART_NUMBER, CLR, AGE",
     "Part 3.5c (AGE)",
     "Refreshed by Contribution / age computation.",
     "Optional",
     "AGE stays NULL — no downstream consumer breaks.",
     "Used for reporting / audit."],
    ["Master_CONT_SZ", "Contribution pipeline (presets)",
     "(ST_CD, MAJ_CAT, SZ)",
     "ST_CD, MAJ_CAT, SZ, CONT",
     "Stage B.2 _enrich_size_cont (priority 1 store, priority 2 CO)",
     "Configure via Contribution UI → Presets.",
     "Yes (or fallback to FNL_Q ratio / uniform)",
     "Falls back to FNL_Q ratio then 1/var_count — workable but suboptimal.",
     "Use ST_CD='CO' for company-default rows."],
    ["MANUAL_DENSITY (or article-level density table)", "Manual UI / data upload",
     "(WERKS, GEN_ART_NUMBER, CLR)",
     "WERKS, GEN_ART_NUMBER, CLR, MANUAL_DENSITY",
     "Part 4c — overrides ACS_D for OPT_MBQ",
     "Maintained via Listing UI manual entry / upload.",
     "Optional",
     "Hot articles can't be forced higher; default ACS_D used.",
     "Per-article-at-store density override."],

    # ─── Master product
    ["vw_master_product", "Underlying master product view",
     "ARTICLE_NUMBER (= VAR_ART) / GEN_ART_NUMBER",
     "ARTICLE_NUMBER, GEN_ART_NUMBER, GEN_ART_DESC, FAB, MACRO_MVGR, MICRO_MVGR, M_VND_CD, RNG_SEG, M_YARN_02, WEAVE_2, ...",
     "Part 3.6 (GEN_ART_DESC), Part 4 pre-resolve, _discover_grid_hierarchy",
     "Refreshed by master-data upload upstream.",
     "Yes",
     "Sec-cap grids that use MP-resolved cols silently drop.",
     "Single source for article attributes."],

    # ─── Hold / pending tracking
    ["ARS_NL_TBL_HOLD_TRACKING", "Auto-maintained by Approve",
     "(WERKS, VAR_ART, SZ)",
     "WERKS, RDC, MAJ_CAT, GEN_ART_NUMBER, CLR, VAR_ART, SZ, OPT_STATUS, LISTED_DATE, HOLD_QTY_INITIAL, HOLD_REM, IS_CLOSED",
     "Part 3.54 (RL_HOLD_QTY); approve_parked HOLD writes",
     "Auto-written on Approve; DDL guard in Part 8.6.",
     "Auto",
     "Carry-over holds invisible to Part 3.6 — articles with prior TBL may misclassify.",
     "Step A decrement on approve; Step B insert; MSA HOLD sync."],
    ["ARS_PEND_ALC", "Auto-maintained by Approve",
     "(WERKS, RDC, MAJ_CAT, GEN_ART_NUMBER, CLR, VAR_ART, SZ, SESSION_ID)",
     "+ ALLOC_QTY, DO_QTY, IS_CLOSED",
     "approve_parked INSERT; apply_pend_alc_delta",
     "Auto-written on Approve.",
     "Auto",
     "—",
     "Pending allocation ledger; revertable."],
    ["ARS_PEND_ALC_OPERATIONS", "Audit log",
     "OP_ID",
     "OP_ID, OP_TYPE (BDC/DO/MANUAL), OP_KEY, PAYLOAD, REVERTED_AT, REVERTED_BY, REVERT_NOTE",
     "Operations Log + Undo UI",
     "Auto-written on every PEND_ALC write/approve/revert.",
     "Auto",
     "—",
     "Source of truth for soft-revert."],

    # ─── Store master
    ["Master_ALC_INPUT_ST_MASTER", "Store master upload",
     "ST_CD UNIQUE",
     "ST_CD, RDC, ST_NAME, ...",
     "Part 1/2 store CROSS JOIN; RDC backfill on HOLD_TRACKING",
     "External upload / SAP integration.",
     "Yes",
     "No stores → empty listing.",
     "Master list of active stores and their home RDC."],
    ["ARS_STORE_BDC_SCHEDULE", "Schedule UI",
     "ST_CD",
     "ST_CD, MON, TUE, WED, THU, FRI, SAT, IS_ACTIVE",
     "BDC generation (downstream of listing approve)",
     "Configure via Schedule Audit UI.",
     "Optional (only for BDC)",
     "—",
     "Per-day store activation for BDC; SUN intentionally absent."],
]

# ---------------------------------------------------------------------------
# Sheet 3 — RUN OUTPUTS (what gets created — verify after run)
# ---------------------------------------------------------------------------
OUTPUT_HDR = ["Table / artifact", "Created by", "Grain", "Purpose", "Lifecycle"]
OUTPUT_ROWS = [
    ["ARS_LISTING", "Part 1 + 2 (Listing build)", "(WERKS, MAJ_CAT, GEN_ART, CLR)",
     "Raw listing universe — every store × every candidate article.", "Recreated every run; not parked."],
    ["ARS_LISTING_WORKING", "Part 7", "(WERKS, MAJ_CAT, GEN_ART, CLR)",
     "Filtered, enriched, gated listing — input to allocator.", "Parked on approve."],
    ["ARS_LISTED_OPT", "Stage A.5", "(WERKS, MAJ_CAT, GEN_ART, CLR)",
     "Only LISTED_FLAG=1 rows — input to Stage B.", "Dropped at end of Part 8."],
    ["ARS_ALLOC_WORKING", "Stage B.1 (+ Stage C result)", "(WERKS, MAJ_CAT, GEN_ART, CLR, VAR_ART, SZ)",
     "Per-size allocation plan — the actual ship list.", "Parked on approve."],
    ["ARS_ALLOC_MAJCAT_QUEUE", "Part 8.3 seed_queue", "(batch_id, maj_cat)",
     "Per-MAJ_CAT progress tracker for the UI dashboard.", "Persists across batches."],
    ["ARS_STORE_RANKING", "Part 6", "(MAJ_CAT, WERKS)",
     "Per-store W_SCORE / ST_RANK used by waterfall tiebreaker.", "Recreated every run."],
    ["ARS_GRID_HIERARCHY", "Part 7", "(WERKS, MAJ_CAT, +per-grid cols)",
     "Managed table of hierarchy columns active grids contribute.", "Auto-maintained."],
    ["ARS_PARKED_ALLOC", "Part 8.4 snapshot_session_to_parked", "Copy of ARS_ALLOC_WORKING",
     "Snapshot tagged with session_id for user review.", "Until approved/rejected."],
    ["ARS_PARKED_LISTING", "Part 8.4", "Copy of ARS_LISTING_WORKING",
     "Snapshot tagged with session_id.", "Until approved/rejected."],
    ["ARS_HISTORICAL_ALLOC", "approve_parked (on Approve)", "Copy of parked + allocation_number",
     "Permanent history of approved allocations.", "Persists indefinitely."],
    ["ARS_HISTORICAL_LISTING", "approve_parked (on Approve)", "Copy of parked listing",
     "Permanent history of approved listings.", "Persists indefinitely."],
]

# ---------------------------------------------------------------------------
# Sheet 4 — PRE-FLIGHT CHECKLIST (printable / shareable)
# ---------------------------------------------------------------------------
CHECK_HDR = ["#", "Step", "Verify", "How to verify (SQL / UI)", "Owner", "Done?"]
CHECK_ROWS = [
    ["1", "Stores master loaded",
     "Master_ALC_INPUT_ST_MASTER has all active stores with valid RDC.",
     "SELECT COUNT(*) FROM Master_ALC_INPUT_ST_MASTER WHERE RDC IS NOT NULL",
     "Master data team", ""],
    ["2", "MSA pipeline run",
     "ARS_MSA_TOTAL, ARS_MSA_GEN_ART, ARS_MSA_VAR_ART all populated for today.",
     "Run MSA Stock Calculation in UI. Check row counts > 0.",
     "Ops", ""],
    ["3", "Grid Builder — MJ family",
     "ARS_GRID_MJ, ARS_GRID_MJ_GEN_ART, ARS_GRID_MJ_VAR_ART rebuilt.",
     "Run Grid Builder → Run All Active.",
     "Ops", ""],
    ["4", "Grid Builder — secondary grids",
     "Each grid you want active is status='Active' in ARS_GRID_BUILDER.",
     "SELECT grid_name, status FROM ARS_GRID_BUILDER ORDER BY seq.",
     "Allocation team", ""],
    ["5", "Contribution pipeline run",
     "ARS_CALC_ST_MAJ_CAT, ARS_CALC_ST_ART, MASTER_GEN_ART_SALE.SAL_PD all current.",
     "Run Contribution Calc Pipeline in UI.",
     "Ops", ""],
    ["6", "Size mix loaded",
     "Master_CONT_SZ has rows for the active MAJ_CATs (preferably ST_CD-specific).",
     "SELECT DISTINCT ST_CD, MAJ_CAT FROM Master_CONT_SZ.",
     "Contribution team", ""],
    ["7", "Master product view fresh",
     "vw_master_product reflects current article catalogue.",
     "SELECT COUNT(*), MAX(<LAST_UPDATED>) FROM vw_master_product.",
     "Master data team", ""],
    ["8", "Merge rules (only if using MERGE_ grids)",
     "ARS_MERGE_RULES has active rules for each MERGE_<X> parent column.",
     "SELECT parent_col, COUNT(*) FROM ARS_MERGE_RULES WHERE is_active=1 GROUP BY parent_col.",
     "Allocation team", ""],
    ["9", "Manual density (optional)",
     "Manual density overrides set for any focus / hot articles.",
     "Listing UI → Manual Density entry. Or query MANUAL_DENSITY table.",
     "Merch", ""],
    ["10", "Prior holds rolled forward",
     "ARS_NL_TBL_HOLD_TRACKING reflects last cycle's TBL holds (auto on Approve).",
     "SELECT COUNT(*) FROM ARS_NL_TBL_HOLD_TRACKING WHERE IS_CLOSED=0.",
     "n/a (auto)", ""],
    ["11", "UI knobs set",
     "Stock%, MinSz, default_acs_d, MJ_REQ growth, RL/TBC caps, PRI strict toggles.",
     "Listing Generation form.",
     "Allocation team", ""],
    ["12", "RDC scope chosen",
     "rdc_mode + values (own / cross / all).",
     "Listing Generation form.",
     "Allocation team", ""],
    ["13", "MAJ_CATs scope chosen",
     "All MAJ_CATs OR a specific subset for a focused run.",
     "Listing Generation form.",
     "Allocation team", ""],
    ["14", "Run engine choice",
     "allocation_mode = pandas (default) unless debugging.",
     "Listing Generation form.",
     "Allocation team", ""],
    ["15", "Parallel workers",
     "parallel_workers ≤ available CPU cores; default 4.",
     "Listing Generation form.",
     "Ops", ""],
    ["16", "Click Generate",
     "Watch ARS_ALLOC_MAJCAT_QUEUE progress / log telemetry.",
     "UI live progress bar.",
     "Allocation team", ""],
    ["17", "Review parked snapshot",
     "ARS_PARKED_ALLOC + ARS_PARKED_LISTING for this session_id.",
     "Listing UI → Parked Runs / Review.",
     "Reviewer / merch", ""],
    ["18", "Approve or Reject",
     "Approve → PEND_ALC + HOLD tracking writes.",
     "Listing UI → Approve button.",
     "Reviewer / merch", ""],
]

# ---------------------------------------------------------------------------
# Sheet 5 — RECENT CHANGES (so users know what shifted)
# ---------------------------------------------------------------------------
CHANGES_HDR = ["Date", "Change", "Affects", "Action required"]
CHANGES_ROWS = [
    ["2026-05-25", "New Part 7 filter: MJ_DISP_Q > 0", "Stores without MAJ_CAT display capacity now drop from working table.",
     "Ensure ARS_GRID_MJ.DISP_Q populated for every active store × MAJ_CAT."],
    ["2026-05-XX", "Inactive grids skipped in PEND_ALC delta", "Manual add / revert / approve no longer touch Inactive grids' PEND_ALC.",
     "Toggle grids Inactive in ARS_GRID_BUILDER if you don't want them updated."],
    ["2026-05-XX", "SZ_MBQ floor-to-1 patch", "Small CONT × small OPT_MBQ no longer rounds to zero.",
     "No action needed; reruns will show more 1-piece allocations on tiny shares."],
    ["2026-05-16", "Fallback path removed", "Rows with ALLOC_FLAG=0 effectively skipped.",
     "Ensure PRI_CT% reaches 100 for desired articles."],
    ["2026-05-11", "RL now requires MSA_FNL_Q > 0", "Articles with adequate stock but no warehouse supply fall to MIX.",
     "No action; expected behaviour."],
]


def main():
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Cover / TOC
    ws = wb.create_sheet("TOC", 0)
    ws["A1"] = "Listing → Allocation — Inputs Master"
    ws["A1"].font = Font(bold=True, size=18)
    ws["A2"] = "Pre-flight checklist + source map for everything Listing/Alloc needs."
    ws["A3"] = "Maintained alongside backend/app/docs/processes/opt_type_classification.md"
    ws["A2"].font = Font(italic=True, color="595959")
    ws["A3"].font = Font(italic=True, color="595959")

    toc = [
        ("UI Inputs", "All knobs the user sets on the Listing Generation form."),
        ("Master Data Tables", "Tables that must be populated by upstream processes."),
        ("Run Outputs", "What gets created (verify after run)."),
        ("Pre-flight Checklist", "Step-by-step verification before clicking Generate."),
        ("Recent Changes", "What changed recently and what to do about it."),
    ]
    ws["A5"] = "Sheets:"
    ws["A5"].font = Font(bold=True)
    for i, (name, desc) in enumerate(toc, start=6):
        ws[f"A{i}"] = name
        ws[f"A{i}"].font = Font(bold=True, color="305496")
        ws[f"B{i}"] = desc
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 90

    # All sheets
    write_sheet(wb, "UI Inputs", UI_HDR, UI_ROWS,
                col_widths=[24, 32, 12, 14, 18, 60, 32, 38, 12])
    write_sheet(wb, "Master Data Tables", DATA_HDR, DATA_ROWS,
                col_widths=[36, 32, 36, 60, 36, 32, 12, 50, 36])
    write_sheet(wb, "Run Outputs", OUTPUT_HDR, OUTPUT_ROWS,
                col_widths=[32, 30, 42, 60, 40])
    write_sheet(wb, "Pre-flight Checklist", CHECK_HDR, CHECK_ROWS,
                col_widths=[4, 32, 60, 60, 24, 8])
    write_sheet(wb, "Recent Changes", CHANGES_HDR, CHANGES_ROWS,
                col_widths=[14, 50, 60, 60])

    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")
    print(f"  Sheets: TOC, UI Inputs ({len(UI_ROWS)}), Master Data ({len(DATA_ROWS)}), "
          f"Run Outputs ({len(OUTPUT_ROWS)}), Pre-flight ({len(CHECK_ROWS)}), "
          f"Recent Changes ({len(CHANGES_ROWS)})")


if __name__ == "__main__":
    main()
